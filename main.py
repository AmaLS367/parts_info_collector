import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm

from clients.llm_client import LLMClient
from config import settings
from promts.generator import generate_prompt
from utils.db_writer import fetch_all, init_db, save_results_bulk
from utils.io import is_processed
from utils.parse import parse_answer

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler("collector.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def format_output_excel(filepath: str, df: pd.DataFrame | None) -> None:
    if df is None or df.empty:
        logger.warning("No data to save to Excel.")
        return

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for col in ws.columns:
        column_idx = col[0].column
        if not isinstance(column_idx, int):
            continue

        max_length = 0
        col_letter = get_column_letter(column_idx)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length * 1.1, 60)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(filepath)
    logger.info(f"Results formatted and saved to {filepath}")


def main() -> None:
    logger.info("Starting AI Data Collector")
    init_db(settings.target_fields)

    try:
        df = pd.read_excel(settings.input_file, sheet_name=settings.sheet_name)
    except Exception as e:
        logger.error(f"Failed to read input file: {e}")
        return

    llm_client = LLMClient()
    buffer: list[tuple[str, ...]] = []

    for start in range(0, len(df), settings.batch_size):
        batch_df = df.iloc[start : start + settings.batch_size]

        for _, row in tqdm(batch_df.iterrows(), total=len(batch_df), desc="Processing batch"):
            item_id = str(row[settings.column_name])

            if is_processed(item_id):
                logger.debug(f"Skipping {item_id} — already in database")
                continue

            prompt = generate_prompt(item_id, settings.item_label, settings.target_fields)
            answer = llm_client.get_answer(prompt)
            parsed = parse_answer(answer, settings.target_fields)

            row_data = (item_id, *[
                parsed.get(f, "Not found")
                for f in settings.target_fields
                if f != settings.column_name
            ])
            buffer.append(row_data)

        if buffer:
            save_results_bulk(buffer, settings.target_fields)
            buffer.clear()

    final_df = fetch_all()
    format_output_excel(settings.output_file, final_df)
    logger.info("Data collection completed successfully")


if __name__ == "__main__":
    main()
